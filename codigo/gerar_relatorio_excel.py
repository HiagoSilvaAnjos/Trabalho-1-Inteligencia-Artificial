import pandas as pd
import os
import sys

def principal():
    # Caminhos
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    CSV_PATH = os.path.join(BASE_DIR, "resultados", "execucoes.csv")
    OUT_PATH = os.path.join(BASE_DIR, "resultados", "relatorio_execucoes.xlsx")

    if not os.path.exists(CSV_PATH):
        print(f"Erro: Arquivo não encontrado - {CSV_PATH}")
        print("Execute o Hill Climbing para gerar os resultados primeiro.")
        sys.exit(1)

    print(f"Lendo dados de {CSV_PATH}...")
    df = pd.read_csv(CSV_PATH)
    
    # Tratamentos básicos de colunas (remoção de espaços, etc)
    df.columns = [col.strip() for col in df.columns]

    print("Gerando relatório Excel...")
    
    # Utilizando pandas com openpyxl engine
    try:
        writer = pd.ExcelWriter(OUT_PATH, engine='openpyxl')
        df.to_excel(writer, sheet_name='Execuções', index=False)
        
        # Acessar workbook e worksheet para formatação
        workbook = writer.book
        worksheet = writer.sheets['Execuções']
        
        # Importações do openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        # Estilos globais
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        border = Border(left=Side(style='thin', color='DDDDDD'), 
                        right=Side(style='thin', color='DDDDDD'), 
                        top=Side(style='thin', color='DDDDDD'), 
                        bottom=Side(style='thin', color='DDDDDD'))
        
        sucesso_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        falha_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        sucesso_font = Font(color="006100")
        falha_font = Font(color="9C0006")
        
        # Encontrar índices de colunas importantes
        headers = [cell.value for cell in worksheet[1]]
        try:
            # Procura por coluna que tenha 'sucesso' no nome ignorando case
            idx_sucesso = next(i for i, h in enumerate(headers) if 'sucesso' in str(h).lower()) + 1
        except StopIteration:
            idx_sucesso = -1

        # Formatar cabeçalho
        for col in range(1, len(headers) + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
        
        # Adicionar Filtro (AutoFilter)
        col_letter_end = get_column_letter(len(headers))
        worksheet.auto_filter.ref = f"A1:{col_letter_end}{len(df) + 1}"

        # Formatar linhas
        for row in range(2, len(df) + 2):
            is_sucesso = False
            if idx_sucesso != -1:
                val = worksheet.cell(row=row, column=idx_sucesso).value
                is_sucesso = str(val).lower() == 'true'

            for col in range(1, len(headers) + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.alignment = center_align
                cell.border = border
                
                # Condicional para sucesso vs falha
                if is_sucesso:
                    cell.fill = sucesso_fill
                    cell.font = sucesso_font
                else:
                    cell.fill = falha_fill
                    cell.font = falha_font

        # Ajustar larguras das colunas
        for col in range(1, len(headers) + 1):
            letter = get_column_letter(col)
            max_len = len(str(headers[col-1])) + 4
            
            for row in range(2, min(100, len(df) + 2)):
                val = worksheet.cell(row=row, column=col).value
                if val:
                    max_len = max(max_len, len(str(val)) + 2)
            
            # Limitar a largura máxima
            worksheet.column_dimensions[letter].width = min(max_len, 45)

        writer.close()
        print(f"[SUCESSO] Relatório Excel gerado com sucesso em: {OUT_PATH}")

    except Exception as e:
        print(f"[ERRO] Falha ao gerar o Excel: {e}")
        print("Verifique se as bibliotecas pandas e openpyxl estão instaladas (pip install pandas openpyxl).")

if __name__ == "__main__":
    principal()
